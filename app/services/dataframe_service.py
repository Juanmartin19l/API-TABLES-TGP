from io import BytesIO
from typing import List, Dict, Optional

import pandas as pd
import numpy as np

COLUMNS = ["Fecha", "Origen", "Monto", "Retenido", "MEP/CTA BNA", "TOTAL"]


def json_to_df(rows: List[Dict]) -> pd.DataFrame:
    df = pd.DataFrame(rows)
    # Ensure columns exist and order
    for c in COLUMNS:
        if c not in df.columns:
            df[c] = np.nan
    df = df[COLUMNS].copy()
    # Normalize empty -> NaN
    df = df.replace({"": np.nan, None: np.nan})
    # Numeric conversion
    num_cols = ["Monto", "Retenido", "MEP/CTA BNA", "TOTAL"]
    for col in num_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    return df


def apply_business_rules(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    # Retenido always negative if present
    df["Retenido"] = df["Retenido"].apply(lambda x: -abs(x) if pd.notna(x) else x)
    # Ensure positivity for numeric columns except Retenido
    for col in ["Monto", "MEP/CTA BNA", "TOTAL"]:
        df[col] = df[col].apply(lambda x: abs(x) if pd.notna(x) else x)
    # Fecha/Origen/Monto remain as-is (Monto not summed)
    return df


def compute_totals_row(df: pd.DataFrame) -> Dict:
    # Use min_count=1 so sum returns NaN if there are no values
    retenido_sum = df["Retenido"].sum(min_count=1)
    mep_sum = df["MEP/CTA BNA"].sum(min_count=1)
    total_sum = df["TOTAL"].sum(min_count=1)

    def nan_to_none(v):
        return None if pd.isna(v) else float(v)

    totals_row = {
        "Fecha": None,
        "Origen": None,
        "Monto": None,
        "Retenido": nan_to_none(retenido_sum),
        "MEP/CTA BNA": nan_to_none(mep_sum),
        "TOTAL": nan_to_none(total_sum),
    }
    return totals_row


def append_totals_row(df: pd.DataFrame, totals_row: Dict) -> pd.DataFrame:
    df_tot = pd.concat([df, pd.DataFrame([totals_row])], ignore_index=True)
    return df_tot


def df_to_json_rows(df: pd.DataFrame) -> List[Dict]:
    out = df.where(pd.notnull(df), None)
    return out.to_dict(orient="records")


# Excel export helpers
def export_df_to_excel_stream(df: pd.DataFrame, sheet_name: str = "Sheet1") -> BytesIO:
    # This function writes the DataFrame to an in-memory Excel workbook using
    # openpyxl, then applies styling: header fill, borders, number formats,
    # auto column widths, freeze header, and highlights the last-row value
    # cells for Retenido, MEP/CTA BNA and TOTAL.
    from io import BytesIO
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
    from pandas import ExcelWriter

    out_buf = BytesIO()

    # 1) Write DataFrame to buffer using openpyxl engine
    with ExcelWriter(out_buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

    # 2) Load workbook for styling
    out_buf.seek(0)
    wb = load_workbook(out_buf)
    ws = wb[sheet_name]

    # Styles
    header_fill = PatternFill(
        start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"
    )
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_side = Side(border_style="thin", color="000000")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    highlight_fill = PatternFill(
        start_color="FFD966", end_color="FFD966", fill_type="solid"
    )

    money_fmt = "#,##0.00"  # no currency symbol

    # Map column names to letters
    col_letters = {}
    for idx, col in enumerate(df.columns, start=1):
        col_letters[col] = get_column_letter(idx)

    # 3) Header styling
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    # 4) Apply number format and borders to monetary columns
    money_columns = ["Monto", "Retenido", "MEP/CTA BNA", "TOTAL"]
    for col in money_columns:
        if col in col_letters:
            col_letter = col_letters[col]
            for row in range(2, ws.max_row + 1):
                cell = ws[f"{col_letter}{row}"]
                cell.number_format = money_fmt
                cell.alignment = right_align
                cell.border = thin_border

    # 5) Apply borders and default alignment for other cells
    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            # If cell is general/unformatted, align left for text
            if getattr(cell, "number_format", "General") == "General":
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # 6) Highlight last row only in the 3 value cells
    last_row = ws.max_row
    highlight_cols = ["Retenido", "MEP/CTA BNA", "TOTAL"]
    for col in highlight_cols:
        if col in col_letters:
            cell = ws[f"{col_letters[col]}{last_row}"]
            cell.fill = highlight_fill
            cell.font = Font(bold=True)

    # 7) Auto-adjust column widths
    for idx, col in enumerate(df.columns, start=1):
        col_letter = get_column_letter(idx)
        values = [str(col)] + [
            "" if v is None else str(v) for v in df[col].fillna("").astype(str).tolist()
        ]
        max_length = max(len(v) for v in values)
        adjusted_width = min(60, max(10, max_length + 2))
        ws.column_dimensions[col_letter].width = adjusted_width

    # 8) Freeze header row
    ws.freeze_panes = ws["A2"]

    # 9) Save workbook to BytesIO and return
    final_buf = BytesIO()
    wb.save(final_buf)
    final_buf.seek(0)
    return final_buf


def export_df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    buf = export_df_to_excel_stream(df, sheet_name)
    return buf.getvalue()
